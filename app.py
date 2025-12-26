# =============================================================
# === 1. KHAI BÁO THƯ VIỆN (IMPORTS) ===
# =============================================================
from flask import Flask, render_template, request, jsonify, Response, redirect, url_for, flash, send_file
from flask_socketio import SocketIO, emit
from flask_migrate import Migrate
from database import db, Class, Student, Subject, Session, SpeechLog, Grade
import os
from database import Class 
from dotenv import load_dotenv
from sqlalchemy import func, desc, text
from datetime import datetime
import pandas as pd
import joblib
import json
import io
import csv
from flask import request, redirect, url_for 
from database import User 
from flask_login import LoginManager, login_user, logout_user, current_user, login_required
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl import Workbook

# =============================================================
# === 2. CẤU HÌNH HỆ THỐNG ===
# =============================================================
load_dotenv()
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'secret')


# Cấu hình Database
db_user = os.getenv('DB_USER', 'root')
db_password = os.getenv('DB_PASSWORD', '')
db_host = os.getenv('DB_HOST', 'localhost')
db_name = os.getenv('DB_NAME', 'pbl4')
app.config['SQLALCHEMY_DATABASE_URI'] = f'mysql+pymysql://{db_user}:{db_password}@{db_host}/{db_name}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
socketio = SocketIO(app, cors_allowed_origins="*")
migrate = Migrate(app, db)

# =============================================================
# === CẤU HÌNH FLASK-LOGIN ===
# =============================================================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login' # Tên HÀM của route đăng nhập
login_manager.login_message = "Vui lòng đăng nhập để truy cập trang này."
login_manager.login_message_category = "info"
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Bạn không có quyền truy cập trang này.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function
@login_manager.user_loader
def load_user(user_id):
    # Flask-Login sẽ dùng hàm này để lấy thông tin user từ ID lưu trong session
    return db.session.get(User, int(user_id))

# Cấu hình hệ số điểm theo quy định giáo viên
HE_SO = {
    'Điểm miệng': 1,
    '15 phút': 1,
    '1 tiết': 2,
    'Học kỳ': 3
}
def generate_ai_comment(subject_name, avg_score, speech_count, class_avg_speech):
    """
    AI Phân tích mối tương quan giữa Điểm số và Dữ liệu Face ID (Phát biểu)
    """
    # 1. Xác định mức độ điểm số
    is_good_grade = avg_score >= 8.0
    is_bad_grade = avg_score < 5.0
    
    # 2. Xác định mức độ năng nổ (Dựa trên trung bình lớp)
    # Giả sử phát biểu > trung bình lớp 20% là năng nổ
    is_active = speech_count > (class_avg_speech * 1.2)
    is_passive = speech_count < (class_avg_speech * 0.6)

    # 3. Phân tích tương quan (Matrix Analysis)
    
    # NHÓM 1: CÔNG TỬ BẠC LIÊU (Giỏi và Năng nổ)
    if is_good_grade and is_active:
        return f"Phong độ xuất sắc. AI ghi nhận sự tương quan hoàn hảo giữa tư duy và thái độ xây dựng bài. Tiếp tục phát huy vai trò dẫn dắt lớp."

    # NHÓM 2: THIÊN TÀI TRẦM LẶNG (Giỏi nhưng thụ động)
    if is_good_grade and is_passive:
        return f"Tiếp thu kiến thức rất tốt nhưng dữ liệu Face ID cho thấy em khá trầm tính. AI khuyến khích em chia sẻ ý kiến nhiều hơn để rèn luyện kỹ năng mềm."

    # NHÓM 3: HỌC TÀI THI PHẬN (Điểm thấp nhưng cực kỳ năng nổ)
    if not is_good_grade and is_active:
        return f"AI đánh giá cao sự nỗ lực vượt bậc qua {speech_count} lượt phát biểu. Tuy nhiên kết quả thi cử chưa tương xứng. Cần rà soát lại phương pháp làm bài hoặc hổng kiến thức căn bản."

    # NHÓM 4: CẦN CHÚ Ý ĐẶC BIỆT (Điểm thấp và thụ động)
    if is_bad_grade and is_passive:
        return f"Cảnh báo: Dữ liệu cho thấy sự thiếu tập trung cả về kiến thức lẫn tương tác. Cần GV bộ môn và phụ huynh sát sao hơn để tìm nguyên nhân mất gốc."

    # NHÓM TRUNG BÌNH (Mặc định)
    if is_active:
        return f"Thái độ học tập tích cực ({speech_count} lần phát biểu). Cần tập trung hơn vào các phần kiến thức khó để nâng mức điểm {avg_score} lên cao hơn."
    
    return f"Hoàn thành mức độ cơ bản của môn học. Cần chủ động hơn trong các tiết học để AI ghi nhận sự tiến bộ về thái độ."
# =============================================================
# === 3. ROUTES GIAO DIỆN (VIEW) ===
# =============================================================
@app.route('/register')
def register():
    # Không cho phép đăng ký công khai nữa
    flash('Chức năng đăng ký công khai đã bị vô hiệu hóa. Vui lòng liên hệ Admin để được cấp tài khoản.', 'info')
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        user = User.query.filter_by(username=username).first()

        if not user or not check_password_hash(user.password_hash, password):
            flash('Tên đăng nhập hoặc mật khẩu không đúng.', 'danger')
            return redirect(url_for('login'))
        
        login_user(user)

        # Chuyển hướng đến trang tiếp theo nếu có, nếu không thì về dashboard
        next_page = request.args.get('next')
        if user.role == 'admin':
            return redirect(next_page or url_for('admin_classes'))
        else:
            return redirect(next_page or url_for('dashboard'))

    return render_template('login.html')


@app.route('/logout')
@login_required 
def logout():
    logout_user()
    flash('Bạn đã đăng xuất.', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required 
def dashboard(): 
    return render_template('dashboard.html')


@app.route('/subjects')
@login_required 
def subjects(): 
    return render_template('subjects.html')

@app.route('/students')
@login_required 
def students(): 
    return render_template('students.html')

@app.route('/stats')
@login_required 
def stats(): 
    return render_template('stats.html') 

@app.route('/live-class')
@login_required 
def liveclass(): 
    return render_template('liveclass.html')

# =============================================================
# === 3. ROUTES GIAO DIỆN ADMIN ===
# =============================================================

@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    classes = Class.query.all()
    return render_template('admin/dashboard.html', classes=classes)

@app.route('/admin/classes', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_classes():
    if request.method == 'POST':
        name = request.form.get('class_name')
        year = request.form.get('academic_year')
        if name and year:
            existing_class = Class.query.filter_by(name=name).first()
            if existing_class:
                flash(f"Lỗi: Tên lớp '{name}' đã tồn tại.", 'danger')
            else:
                new_class = Class(name=name, academic_year=year)
                db.session.add(new_class)
                db.session.commit()
                flash(f"Đã thêm thành công lớp học '{name}'.", 'success')
            return redirect(url_for('admin_classes'))
    all_classes = Class.query.order_by(Class.name).all()
    return render_template('admin/classes.html', classes=all_classes)

@app.route('/admin/users', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_users():
    if request.method == 'POST':
        # Logic tạo tài khoản giáo viên
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        new_teacher = User(username=username, email=email, password_hash=hashed_password, role='teacher')
        db.session.add(new_teacher)
        db.session.commit()
        flash(f"Đã tạo tài khoản cho giáo viên {username}", "success")
        return redirect(url_for('admin_users'))

    # Lấy dữ liệu cho 2 bảng
    classes = Class.query.order_by(Class.name).all()
    teachers = User.query.filter_by(role='teacher').order_by(User.username).all()
    return render_template('admin/users.html', classes=classes, teachers=teachers)

# --- API PHÂN CÔNG GIÁO VIÊN VÀO LỚP (Dựa trên ID Lớp) ---
@app.route('/api/classes/<int:class_id>/assign_teacher', methods=['POST'])
@login_required
@admin_required
def assign_teacher_to_class(class_id):
    data = request.get_json()
    new_teacher_id = data.get('teacher_id')
    confirm_switch = data.get('confirm_switch', False) # Cờ xác nhận từ popup

    # 1. Trường hợp bỏ trống giáo viên cho lớp
    if not new_teacher_id or int(new_teacher_id) == 0:
        old_teacher = User.query.filter_by(class_id=class_id).first()
        if old_teacher:
            old_teacher.class_id = None
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Đã để trống giáo viên cho lớp này'})

    teacher = db.session.get(User, new_teacher_id)
    if not teacher:
        return jsonify({'status': 'error', 'message': 'Không tìm thấy giáo viên'}), 404

    # 2. KIỂM TRA XUNG ĐỘT: Giáo viên này đã quản lý lớp KHÁC chưa?
    # Tìm lớp mà giáo viên này đang giữ (ngoại trừ lớp hiện tại đang xét)
    current_class_of_teacher = Class.query.filter(Class.id != class_id).join(User).filter(User.id == new_teacher_id).first()

    if current_class_of_teacher and not confirm_switch:
        return jsonify({
            'status': 'conflict',
            'message': f'Giáo viên "{teacher.username}" hiện đang quản lý lớp "{current_class_of_teacher.name}". Bạn có muốn đổi không?'
        }), 409 # Gửi mã lỗi 409 để JS nhận diện

    # 3. THỰC HIỆN ĐỔI (Nếu không trùng hoặc đã nhấn "OK" ở popup)
    
    # Gỡ giáo viên cũ đang giữ lớp này ra (nếu có)
    old_teacher_of_class = User.query.filter_by(class_id=class_id).first()
    if old_teacher_of_class:
        old_teacher_of_class.class_id = None

    # Gỡ giáo viên mới ra khỏi lớp cũ của họ (nếu có)
    teacher.class_id = class_id
    
    db.session.commit()
    return jsonify({'status': 'success', 'message': 'Cập nhật phân công thành công'})

# --- API XÓA GIÁO VIÊN ---
@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_teacher(user_id):
    user = db.session.get(User, user_id)
    if not user or user.role == 'admin':
        return jsonify({'status': 'error', 'message': 'Không thể xóa'}), 400
    
    db.session.delete(user)
    db.session.commit()
    return jsonify({'status': 'success', 'message': 'Đã xóa giáo viên'})

@app.route('/admin/students')
@admin_required 
def admin_students():
    return render_template('admin/students.html')

@app.route('/admin/subjects')
@admin_required 
def admin_subjects():
    return render_template('admin/subjects.html')

# === 4. API ENDPOINTS (LOGIC) ===
# =============================================================

# --- API NHẬN DIỆN TỪ AI SERVER ---
@app.route('/api/recognize', methods=['POST'])
def recognize_api():
    data = request.get_json()
    student_code = data.get('student_code')
    
    current_session = Session.query.filter_by(status='ongoing').first()
    if not current_session:
        return jsonify({"status": "error", "message": "Chưa bắt đầu buổi học"}), 400
    
    student = Student.query.filter_by(student_code=student_code).first()
    if not student:
        return jsonify({"status": "error", "message": "Không tìm thấy HS"}), 404

    socketio.emit('pending_recognition', {
        'student_code': student.student_code,
        'full_name': student.full_name,
        'student_id': student.id
    }, namespace='/live')
    
    return jsonify({"status": "pending"})
# --- QUẢN LÝ MÔN HỌC ---
@app.route('/api/subjects', methods=['GET', 'POST'])
@login_required
def subjects_api():
    user_class_id = current_user.class_id
    if not user_class_id:
        return jsonify({"message": "Bạn chưa được phân công lớp"}), 403

    if request.method == 'POST':
        data = request.get_json()
        new_subject = Subject(
            name=data['name'], 
            icon=data.get('icon', '📚'), 
            category=data.get('category'), 
            class_id=user_class_id # Thay 1 bằng user_class_id
        )
        db.session.add(new_subject)
        db.session.commit()
        return jsonify({'status': 'success', 'message': f"Đã thêm môn học '{data['name']}'."})
    
    # Chỉ lấy môn học của lớp mình
    subjects = Subject.query.filter_by(class_id=user_class_id).all()
    result = []
    for s in subjects:
        session_count = Session.query.filter_by(subject_id=s.id).count()
        total_speeches = SpeechLog.query.join(Session).filter(Session.subject_id == s.id).count()
        result.append({'id': s.id, 'name': s.name, 'icon': s.icon, 'category': s.category, 'session_count': session_count, 'total_speeches': total_speeches})
    return jsonify(result)

@app.route('/api/subjects/<int:subject_id>', methods=['GET', 'PUT', 'DELETE'])
def single_subject_api(subject_id):
    subject = db.session.get(Subject, subject_id)
    if not subject: 
        return jsonify({'status': 'error', 'message': 'Không tìm thấy môn học'}), 404

    if request.method == 'PUT':
        data = request.get_json()
        subject.name = data['name']
        subject.icon = data.get('icon', subject.icon)
        subject.category = data.get('category', subject.category)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Cập nhật thành công'})
    
    elif request.method == 'DELETE':
        try:
            Grade.query.filter_by(subject_id=subject.id).delete()
            sessions = Session.query.filter_by(subject_id=subject.id).all()
            session_ids = [s.id for s in sessions]
            if session_ids:
                SpeechLog.query.filter(SpeechLog.session_id.in_(session_ids)).delete(synchronize_session=False)
            Session.query.filter_by(subject_id=subject.id).delete()
            db.session.delete(subject)
            db.session.commit()
            return jsonify({'status': 'success', 'message': f"Đã xóa môn học '{subject.name}' và toàn bộ dữ liệu liên quan."})
        except Exception as e:
            db.session.rollback()
            print(f"Lỗi khi xóa môn học: {str(e)}")
            return jsonify({'status': 'error', 'message': f'Không thể xóa môn học do có lỗi ràng buộc dữ liệu.'}), 500
    
    return jsonify({'id': subject.id, 'name': subject.name, 'icon': subject.icon, 'category': subject.category})

# --- QUẢN LÝ HỌC SINH ---
@app.route('/api/students', methods=['GET', 'POST'])
@login_required
def students_api():
    user_class_id = current_user.class_id
    if not user_class_id:
        return jsonify([]), 403

    if request.method == 'POST':
        data = request.get_json()
        dob = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date() if data.get('date_of_birth') else None
        new_student = Student(
            full_name=data['name'], 
            student_code=data['code'], 
            class_id=user_class_id, # Thay 1 bằng user_class_id
            date_of_birth=dob, 
            gender=data.get('gender')
        )
        db.session.add(new_student)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Thêm học sinh thành công'})
    
    # Chỉ lấy học sinh của lớp mình
    students = Student.query.filter_by(class_id=user_class_id).order_by(Student.full_name).all()
    return jsonify([{'id': s.id, 'full_name': s.full_name, 'student_code': s.student_code} for s in students])

@app.route('/api/students/<int:student_id>', methods=['GET', 'PUT', 'DELETE'])
def single_student_api(student_id):
    student = db.session.get(Student, student_id)
    if not student: return jsonify({'status': 'error', 'message': 'Không tìm thấy học sinh'}), 404

    if request.method == 'PUT':
        data = request.get_json()
        student.full_name = data.get('name', student.full_name)
        if data.get('date_of_birth'):
            student.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
        student.gender = data.get('gender', student.gender)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Cập nhật thành công'})
    
    elif request.method == 'DELETE':
        SpeechLog.query.filter_by(student_id=student.id).delete()
        Grade.query.filter_by(student_id=student.id).delete()
        db.session.delete(student)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Đã xóa học sinh'})
    
    return jsonify({
        'id': student.id, 
        'full_name': student.full_name, 
        'student_code': student.student_code, 
        'date_of_birth': student.date_of_birth.isoformat() if student.date_of_birth else None, 
        'gender': student.gender
    })

# --- QUẢN LÝ ĐIỂM SỐ ---
@app.route('/api/students/<int:student_id>/grades', methods=['GET'])
def get_grades(student_id):
    grades = Grade.query.filter_by(student_id=student_id).all()
    return jsonify([{
        'id': g.id, 
        'score': g.score, 
        'grade_type': g.grade_type, 
        'term': g.term, 
        'exam_date': g.exam_date.isoformat(), 
        'subject_id': g.subject_id, 
        'subject_name': g.subject.name
    } for g in grades])
    
@app.route('/api/students/<int:student_id>/export_smart_report')
@login_required
def export_smart_report(student_id):
    student = db.session.get(Student, student_id)
    if not student: return "Học sinh không tồn tại", 404

    # Tính trung bình phát biểu của cả lớp để AI lấy làm mốc so sánh
    total_speeches_class = SpeechLog.query.count()
    total_students_class = Student.query.count()
    class_avg_speech = total_speeches_class / (total_students_class or 1)

    subjects = Subject.query.filter_by(class_id=student.class_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Report Card"

    # Style Header chuyên nghiệp
    fill_blue = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
    top=Side(style='thin'), bottom=Side(style='thin'))

    # Vẽ Tiêu đề
    ws.merge_cells('A1:H1')
    ws['A1'] = f"BÁO CÁO HỌC TẬP THÔNG MINH (DỰA TRÊN AI & NHẬN DIỆN KHUÔN MẶT)"
    ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f"Học sinh: {student.full_name}"
    ws['A3'] = f"Mã FaceID: {student.student_code}"
    ws['E2'] = f"Lớp: {student.class_info.name}"
    ws['E3'] = f"Trung bình phát biểu lớp: {round(class_avg_speech, 1)} lần/kỳ"

    # Tạo Header bảng
    headers = ["STT", "Môn Học", "Điểm TP", "Thi HK", "TB Môn", "Hệ số", "Phát biểu (FaceID)", "Phân tích từ AI"]
    ws.append([])
    ws.append(headers)
    for cell in ws[5]:
        cell.fill = fill_blue
        cell.font = font_white
        cell.alignment = Alignment(horizontal='center')

    # Đổ dữ liệu
    row_idx = 6
    for i, sub in enumerate(subjects):
        grades = Grade.query.filter_by(student_id=student_id, subject_id=sub.id).all()
        # Lấy số lần phát biểu từ nhật ký Face ID
        speech_count = SpeechLog.query.join(Session).filter(
            SpeechLog.student_id == student_id, Session.subject_id == sub.id
        ).count()

        # Tính toán điểm
        tu_so, mau_so = 0, 0
        tp_list = []
        hk_score = ""
        for g in grades:
            w = HE_SO.get(g.grade_type, 1)
            tu_so += g.score * w
            mau_so += w
            if g.grade_type == 'Học kỳ': hk_score = g.score
            else: tp_list.append(str(g.score))

        avg_sub = round(tu_so / mau_so, 1) if mau_so > 0 else 0
        ai_msg = generate_ai_comment(sub.name, avg_sub, speech_count, class_avg_speech)

        ws.append([i+1, sub.name, ", ".join(tp_list), hk_score, avg_sub, "Theo quy định", f"{speech_count} lần", ai_msg])
        
        # Kẻ bảng
        for cell in ws[row_idx]:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        row_idx += 1

    # Điều chỉnh độ rộng cột
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 60 # Cột nhận xét AI cần rộng

    # Xuất file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    as_attachment=True, download_name=f"BaoCaoAI_{student.student_code}.xlsx")

@app.route('/api/grades', methods=['POST'])
def add_grade():
    data = request.get_json()
    try:
        exam_date = datetime.strptime(data['exam_date'], '%Y-%m-%d').date() if data.get('exam_date') else datetime.utcnow().date()
        new_grade = Grade(student_id=data['student_id'], subject_id=data['subject_id'], score=float(data['score']), grade_type=data.get('grade_type'), term=data.get('term'), exam_date=exam_date)
        db.session.add(new_grade)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Đã thêm điểm.'}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'Lỗi server: {e}'}), 500

@app.route('/api/grades/<int:grade_id>', methods=['PUT', 'DELETE'])
def update_delete_grade(grade_id):
    grade = db.session.get(Grade, grade_id)
    if not grade: return jsonify({'status': 'error', 'message': 'Không tìm thấy điểm'}), 404
    
    if request.method == 'DELETE':
        db.session.delete(grade)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Đã xóa điểm'})
    
    if request.method == 'PUT':
        data = request.get_json()
        grade.score = float(data['score'])
        grade.grade_type = data['grade_type']
        grade.term = data['term']
        grade.subject_id = data['subject_id']
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Đã cập nhật điểm'})
# --- BUỔI HỌC (SESSION) ---
@app.route('/api/sessions/start', methods=['POST'])
@login_required
def start_session_api():
    user_class_id = current_user.class_id
    data = request.get_json()
    subject_id = data.get('subject_id')
    
    # Kiểm tra xem môn học đó có thuộc lớp của giáo viên này không
    subject = Subject.query.filter_by(id=subject_id, class_id=user_class_id).first()
    if not subject:
        return jsonify({"status": "error", "message": "Môn học không hợp lệ cho lớp này"}), 403

    ongoing = Session.query.filter_by(status='ongoing', class_id=user_class_id).first()
    if ongoing: 
        return jsonify({"status": "error", "message": "Một buổi học khác của lớp đang diễn ra."}), 409

    new_session = Session(subject_id=subject_id, status='ongoing', class_id=user_class_id) # Thay 1
    db.session.add(new_session)
    db.session.commit()
    return jsonify({"status": "success", "session_id": new_session.id})

@app.route('/api/sessions/end', methods=['POST'])
def end_session_api():
    data = request.get_json()
    session = db.session.get(Session, data.get('session_id'))
    if not session: return jsonify({"status": "error", "message": "Không tìm thấy buổi học"}), 404
    
    session.status = 'ended'
    session.end_time = datetime.utcnow()
    session.speech_count = SpeechLog.query.filter_by(session_id=session.id).count()
    db.session.commit()
    return jsonify({"status": "success", "message": "Buổi học đã kết thúc."})

@app.route('/api/sessions/current', methods=['GET'])
def get_current_session_api():
    session = Session.query.filter_by(status='ongoing').first()
    if session:
        counts = db.session.query(Student.student_code, func.count(SpeechLog.id)).join(SpeechLog).filter(SpeechLog.session_id == session.id).group_by(Student.student_code).all()
        return jsonify({"status": "found", "session_id": session.id, "subject_name": session.subject.name, "speech_counts": dict(counts)})
    return jsonify({"status": "not_found"})

# --- LỊCH SỬ VÀ CHI TIẾT ---
@app.route('/api/subjects/<int:subject_id>/history')
def session_history(subject_id):
    sessions = Session.query.filter_by(subject_id=subject_id).order_by(Session.id.desc()).all()
    res = []
    for i, s in enumerate(sessions):
        count = SpeechLog.query.filter_by(session_id=s.id).count()
        res.append({
            "id": s.id, 
            "session_number": len(sessions) - i, 
            "end_time": s.end_time.strftime('%d/%m/%Y %H:%M') if s.end_time else "Đang dạy", 
            "speech_count": count
        })
    return jsonify(res)

@app.route('/api/sessions/<int:session_id>/details')
def session_details_api(session_id):
    details = db.session.query(Student.full_name, func.count(SpeechLog.id)).join(SpeechLog).filter(SpeechLog.session_id == session_id).group_by(Student.id).all()
    return jsonify([{"name": d[0], "count": d[1]} for d in details])
# --- THỐNG KÊ DASHBOARD ---
@app.route('/api/dashboard_stats')
@login_required
def dashboard_stats_api():
    user_class_id = current_user.class_id
    if not user_class_id:
        return jsonify({"message": "Chưa có dữ liệu lớp"}), 403
        
    try:
        # 1. KPI tổng quan
        stats = {
            'subjects': Subject.query.filter_by(class_id=user_class_id).count(),
            'students': Student.query.filter_by(class_id=user_class_id).count(),
            'sessions': Session.query.filter_by(class_id=user_class_id).count(),
            'speeches': SpeechLog.query.join(Session).filter(Session.class_id == user_class_id).count()
        }
        
        recent = []
        # 2. Lấy 3 buổi học có THỜI GIAN KẾT THÚC mới nhất (Đưa cái vừa học xong lên đầu)
        # Sắp xếp theo end_time giảm dần (desc)
        sessions = Session.query.filter_by(class_id=user_class_id)\
                          .order_by(desc(Session.end_time), desc(Session.id))\
                          .limit(3).all()

        for s in sessions:
            # 3. LOGIC QUAN TRỌNG: Tính số hiệu (#) dựa trên RANK thời gian
            # Một buổi là "Buổi thứ bao nhiêu" = Tổng số buổi có thời gian kết thúc <= nó
            real_session_number = Session.query.filter(
                Session.class_id == user_class_id,
                db.or_(
                    Session.end_time < s.end_time,
                    db.and_(Session.end_time == s.end_time, Session.id <= s.id)
                )
            ).count()

            recent.append({
                'subject_name': s.subject.name,
                'session_number': real_session_number,
                # Đây là thời gian KẾT THÚC buổi học
                'end_time': s.end_time.strftime('%d/%m/%Y %H:%M') if s.end_time else "Đang diễn ra",
                'speech_count': SpeechLog.query.filter_by(session_id=s.id).count()
            })
            
        return jsonify({'stats': stats, 'recent_activity': recent})
    except Exception as e: 
        print(f"Lỗi Dashboard: {e}")
        return jsonify({"message": f"Lỗi: {str(e)}"}), 500
    
@app.route('/api/untrained_faces')
def untrained_faces_api():
    DATASET_PATH = "datasets/faces"
    try:
        trained = [d for d in os.listdir(DATASET_PATH) if os.path.isdir(os.path.join(DATASET_PATH, d))]
        used = [s.student_code for s in Student.query.all()]
        return jsonify([c for c in trained if c not in used])
    except: return jsonify([])
# --- THỐNG KÊ CHI TIẾT (STATS TAB) ---
@app.route('/api/statistics')
@login_required
def statistics_api():
    try:
        user_class_id = current_user.class_id


        # 1. Bảng xếp hạng học sinh trong lớp 
        all_students_ranking_query = db.session.query(
            Student.id, Student.full_name, func.count(SpeechLog.id).label('total_speeches')
        ).outerjoin(SpeechLog).filter(Student.class_id == user_class_id)\
         .group_by(Student.id).order_by(desc('total_speeches')).all()
       
        all_students_ranking = [{'id': s[0], 'name': s[1], 'speeches': s[2]} for s in all_students_ranking_query]


        # 2. KPI tổng quan
        total_sessions = Session.query.filter_by(class_id=user_class_id).count()
        total_speeches = SpeechLog.query.join(Session).filter(Session.class_id == user_class_id).count()
        total_students = Student.query.filter_by(class_id=user_class_id).count()
        most_active_student = all_students_ranking[0]['name'] if all_students_ranking else "N/A"


        # 3. Phân tích chi tiết từng môn học
        subjects = Subject.query.filter_by(class_id=user_class_id).all()
        subject_analysis = []
        for s in subjects:
            # Truy vấn tìm học sinh tích cực nhất cho môn học s này
            top_stu_query = db.session.query(
                Student.full_name, func.count(SpeechLog.id).label('count')
            ).join(SpeechLog, Student.id == SpeechLog.student_id)\
             .join(Session, SpeechLog.session_id == Session.id)\
             .filter(Session.subject_id == s.id)\
             .group_by(Student.id)\
             .order_by(desc('count')).first()


            if top_stu_query:
                top_name = top_stu_query[0]
                top_count = top_stu_query[1]
            else:
                top_name = "N/A"
                top_count = 0


            subject_analysis.append({
                'id': s.id,
                'name': s.name,
                'icon': s.icon,
                'session_count': Session.query.filter_by(subject_id=s.id).count(),
                'total_speeches': db.session.query(func.count(SpeechLog.id)).join(Session).filter(Session.subject_id == s.id).scalar() or 0,
                'top_student_name': top_name,
                'top_student_speeches': top_count
            })
       
        return jsonify({
            'kpis': {
                "total_sessions": total_sessions, "total_speeches": total_speeches,
                "total_students": total_students, "most_active_student": most_active_student
            },
            'all_students_ranking': all_students_ranking,
            'subject_analysis': subject_analysis
        })
    except Exception as e:
        print(f"Lỗi Statistics API: {e}")
        return jsonify({"message": str(e)}), 500
@app.route('/api/students/<int:student_id>/analysis')
@login_required
def analyze_student_api(student_id):
    try:
        # 1. Kiểm tra quyền truy cập (Học sinh phải thuộc lớp của giáo viên đang đăng nhập)
        user_class_id = current_user.class_id
        student = Student.query.filter_by(id=student_id, class_id=user_class_id).first()
        if not student:
            return jsonify({"message": "Học sinh không tồn tại hoặc không thuộc lớp của bạn"}), 404

        # 2. Khai báo các nhóm môn học cần phân tích
        categories = ['Khoa học Tự nhiên', 'Khoa học Xã hội', 'Ngoại ngữ', 'Năng khiếu']
        radar_values = []
        raw_scores = {} # Lưu điểm trung bình thực tế để hiển thị minh chứng

        for cat in categories:
            # Truy vấn điểm trung bình thực tế của nhóm môn này
            g_avg = db.session.query(func.avg(Grade.score)).join(Subject).filter(
                Grade.student_id == student_id, 
                Subject.category == cat
            ).scalar() or 0
            
            # Lưu lại con số thực tế (ví dụ: 8.5)
            raw_scores[cat] = round(float(g_avg), 1)

            # Truy vấn số lượt phát biểu (dữ liệu từ Face ID)
            s_cnt = db.session.query(func.count(SpeechLog.id)).join(Session).join(Subject).filter(
                SpeechLog.student_id == student_id, 
                Subject.category == cat
            ).count()
           
            # Tính toán giá trị Radar (Điểm năng lực tổng hợp)
            # Công thức: 70% từ điểm số + 30% từ thái độ (tối đa 15 lượt phát biểu được tính điểm 10)
            attitude_score = min(s_cnt, 15) * (10 / 15)
            radar_val = (float(g_avg) * 0.7) + (attitude_score * 0.3)
            radar_values.append(round(radar_val, 1))

        # 3. Xác định nhóm môn có điểm Radar cao nhất (Thế mạnh thực tế)
        max_radar_val = max(radar_values)
        
        if max_radar_val == 0:
            # Nếu điểm Radar cao nhất vẫn là 0 -> Chưa có dữ liệu gì cả
            actual_best = "Chưa thể xác định"
            best_score_text = "0.0" # Điểm Radar minh chứng
        else:
            # Tìm danh sách các môn có điểm RADAR bằng mức cao nhất
            best_indices = [i for i, val in enumerate(radar_values) if val == max_radar_val]
            
            if len(best_indices) == len(categories):
                actual_best = "Phát triển toàn diện"
            elif len(best_indices) > 1:
                # Nếu có 2 môn bằng điểm Radar nhau (Ví dụ: KHTN 8.5 và Ngoại ngữ 8.5)
                actual_best = " & ".join([categories[i] for i in best_indices])
            else:
                # Một môn duy nhất có điểm Radar cao nhất
                actual_best = categories[best_indices[0]]
            
            # Hiển thị ĐIỂM RADAR làm minh chứng thay vì điểm trung bình
            best_score_text = str(max_radar_val)

        # 4. Lấy nhận định phong cách từ kết quả huấn luyện AI (file JSON)
        try:
            # File này được tạo ra từ script analyze_trends.py
            with open('behavioral_analysis.json', 'r', encoding='utf-8') as f:
                behavior_data = json.load(f)
            ai_style = behavior_data.get(str(student_id), {
                "style": "Đang theo dõi",
                "advice": "Tiếp tục thu thập dữ liệu để AI đưa ra nhận định chính xác hơn."
            })
        except FileNotFoundError:
            ai_style = {
                "style": "Chưa có phân tích AI", 
                "advice": "Hãy yêu cầu Admin chạy huấn luyện AI để cập nhật phong cách học tập."
            }

        # 5. Lấy tổng số lần phát biểu toàn thời gian
        total_speeches = SpeechLog.query.filter_by(student_id=student_id).count()

        # 6. Tổng hợp và trả về kết quả
        return jsonify({
            "kpis": {
                "rank": ai_style['style'],           # Phong cách học tập (ví dụ: Học sinh Năng động)
                "total_speeches": total_speeches,    # Tổng lượt Face ID nhận diện phát biểu
                "best_subject": actual_best          # Nhóm môn thế mạnh nhất
            },
            "raw_scores": raw_scores,                # Gửi thêm điểm số thực tế để minh chứng
            "radar": {
                "labels": categories, 
                "data": radar_values
            },
            "trend": {
                "labels": ["Tuần 1", "Tuần 2", "Tuần 3", "Tuần 4"], 
                "data": [0, 2, 5, total_speeches]    # Giả lập xu hướng, bạn có thể viết query thực nếu cần
            },
            "insight": {
                "tendency": actual_best,
                "reason": f"AI nhận diện em có thế mạnh vượt trội ở nhóm môn **{actual_best}** với điểm trung bình thực tế là **{best_score_text}**. " \
                          f"Kết hợp với dữ liệu hành vi, em được xếp vào nhóm **{ai_style['style']}**. " \
                          f"Lời khuyên: {ai_style['advice']}"
            }
        })

    except Exception as e:
        print(f"Lỗi phân tích học sinh {student_id}: {str(e)}")
        return jsonify({"message": f"Lỗi hệ thống: {str(e)}"}), 500
# --- XUẤT DỮ LIỆU ---
@app.route('/api/sessions/export', methods=['POST'])
def export_session_data():
    try:
        export_data_str = request.form.get('export_data')
        if not export_data_str: return "Không có dữ liệu", 400
        
        data = json.loads(export_data_str)
        stats = data.get('stats', {})
        students = data.get('students', [])
        subject_name = data.get('subject_name', 'Unknown Subject')
        session_id = data.get('session_id', 'Unknown_Session')

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Mã Sinh Viên', 'Họ và Tên', 'Số lần phát biểu'])

        student_map = {s['student_code']: s['full_name'] for s in students}
        for student_code, count in stats.items():
            full_name = student_map.get(student_code, 'Không rõ tên')
            writer.writerow([student_code, full_name, count])

        output.seek(0)
        filename = f"ThongKe_BuoiHoc_{session_id}_{subject_name.replace(' ', '_')}.csv"
        return Response(output, mimetype="text/csv", headers={"Content-Disposition": f"attachment;filename={filename}"})
    except Exception as e:
        print(f"Lỗi khi xuất dữ liệu: {e}")
        return "Có lỗi xảy ra trong quá trình xuất file.", 500
# API để lấy thông tin chi tiết của một lớp (dùng cho form Sửa)
@app.route('/api/classes/<int:class_id>', methods=['GET'])
def get_class_details(class_id):
    class_obj = db.session.get(Class, class_id)
    if class_obj:
        return jsonify({
            'id': class_obj.id,
            'name': class_obj.name,
            'academic_year': class_obj.academic_year
        })
    return jsonify({'error': 'Không tìm thấy lớp học'}), 404

# API để cập nhật thông tin một lớp (dùng cho form Sửa)
@app.route('/api/classes/<int:class_id>', methods=['PUT'])
def update_class(class_id):
    class_obj = db.session.get(Class, class_id)
    if not class_obj:
        return jsonify({'status': 'error', 'message': 'Không tìm thấy lớp học'}), 404
    
    data = request.get_json()
    name = data.get('name')
    year = data.get('academic_year')

    if not name or not year:
        return jsonify({'status': 'error', 'message': 'Tên lớp và năm học không được để trống'}), 400

    # Kiểm tra tên mới có trùng với lớp khác không
    existing_class = Class.query.filter(Class.name == name, Class.id != class_id).first()
    if existing_class:
        return jsonify({'status': 'error', 'message': f"Tên lớp '{name}' đã tồn tại"}), 409

    class_obj.name = name
    class_obj.academic_year = year
    db.session.commit()
    return jsonify({'status': 'success', 'message': 'Cập nhật lớp học thành công'})

# API để xóa một lớp học
@app.route('/api/classes/<int:class_id>', methods=['DELETE'])
def delete_class(class_id):
    class_obj = db.session.get(Class, class_id)
    if not class_obj:
        return jsonify({'status': 'error', 'message': 'Không tìm thấy lớp học'}), 404
    
    # Kiểm tra xem lớp có học sinh hoặc môn học nào không
    if class_obj.students or class_obj.subjects:
        return jsonify({'status': 'error', 'message': 'Không thể xóa lớp học này vì vẫn còn học sinh hoặc môn học.'}), 409

    db.session.delete(class_obj)
    db.session.commit()
    return jsonify({'status': 'success', 'message': 'Đã xóa lớp học thành công'})

# Trong app.py, khu vực API ENDPOINTS
@app.route('/api/users/<int:user_id>/assign_class', methods=['POST'])
@login_required
@admin_required
def assign_class_to_user(user_id):
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({'status': 'error', 'message': 'Không tìm thấy người dùng'}), 404
   
    data = request.get_json()
    class_id = data.get('class_id')

    # Nếu chọn "Chưa phân công"
    if not class_id or int(class_id) == 0:
        user.class_id = None
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Đã hủy phân công.'})

    # KIỂM TRA DUY NHẤT: 1 lớp - 1 giáo viên
    existing_teacher = User.query.filter(User.class_id == class_id, User.id != user_id).first()
    if existing_teacher:
        return jsonify({
            'status': 'error', 
            'message': f'Lớp này đã được phân công cho giáo viên "{existing_teacher.username}".'
        }), 400 

    user.class_id = class_id
    db.session.commit()
    return jsonify({'status': 'success', 'message': 'Cập nhật phân công thành công.'})
# --- API THỐNG KÊ TỔNG QUAN CHO ADMIN ---
# 1. Cập nhật API thống kê tổng quát
@app.route('/api/admin/stats')
@login_required
@admin_required
def admin_stats_api():
    try:
        stats = {
            'teachers': User.query.filter_by(role='teacher').count(),
            'classes': Class.query.count(),
            'students': Student.query.count(),
            'speeches': SpeechLog.query.count()
        }

        # BIỂU ĐỒ 1: Đường đua giữa các lớp (Xếp hạng tổng lượt tương tác)
        race_query = db.session.query(
            Class.name, func.count(SpeechLog.id).label('total')
        ).join(Student, Class.id == Student.class_id)\
         .join(SpeechLog, Student.id == SpeechLog.student_id)\
         .group_by(Class.id).order_by(desc('total')).all()
        
        race_data = {
            'labels': [r[0] for r in race_query],
            'values': [r[1] for r in race_query]
        }

        return jsonify({'kpis': stats, 'race': race_data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# 2. Tạo API mới: Phân tích thế mạnh môn học của một lớp cụ thể
@app.route('/api/admin/class_performance/<int:class_id>')
@login_required
@admin_required
def class_performance_api(class_id):
    # Lấy phân bổ lượt phát biểu theo môn học của lớp được chọn
    performance = db.session.query(
        Subject.name, func.count(SpeechLog.id).label('count')
    ).join(Session, Subject.id == Session.subject_id)\
     .join(SpeechLog, Session.id == SpeechLog.session_id)\
     .join(Student, SpeechLog.student_id == Student.id)\
     .filter(Student.class_id == class_id)\
     .group_by(Subject.id).all()
    
    return jsonify({
        'labels': [p[0] for p in performance] if performance else ["Chưa có dữ liệu"],
        'values': [p[1] for p in performance] if performance else [0]
    })
# =============================================================
# === 5. SOCKET.IO ===
# =============================================================
@socketio.on('confirm_recognition', namespace='/live')
def handle_confirm(data):
    if data.get('action') == 'accept':
        sess = Session.query.filter_by(status='ongoing').first()
        stu = db.session.get(Student, data.get('student_id'))
        if sess and stu:
            db.session.add(SpeechLog(student_id=stu.id, session_id=sess.id))
            db.session.commit()
            socketio.emit('speech_update', {'student_code': stu.student_code, 'full_name': stu.full_name}, namespace='/live')

# =============================================================
# === 6. KHỐI KHỞI ĐỘNG SERVER (PHẢI LUÔN NẰM CUỐI CÙNG) ===
# =============================================================

@app.context_processor
def inject_user_info():
    if current_user.is_authenticated:
        # Lấy thông tin lớp học từ relationship assigned_class (đã định nghĩa trong database.py)
        user_class = current_user.assigned_class
        class_name = user_class.name if user_class else "Chưa phân công"
        
        # Đếm sĩ số học sinh của riêng lớp đó
        total_students = 0
        if user_class:
            total_students = Student.query.filter_by(class_id=user_class.id).count()
            
        return {
            'header_class_name': class_name,
            'header_total_students': total_students,
            'header_username': current_user.username
        }
    return {
        'header_class_name': "N/A",
        'header_total_students': 0,
        'header_username': "Guest"
    }
if __name__ == '__main__':
    with app.app_context():
        # 1. Tự động tạo bảng nếu chưa có
        db.create_all()

        # 2. Tạo lớp mặc định nếu chưa có
        if not db.session.get(Class, 1):
            db.session.add(Class(id=1, name="9/1", academic_year="2025-2026"))
            db.session.commit()
            print("Đã tạo lớp học mặc định 9/1")

        # 3. TẠO TÀI KHOẢN ADMIN MẶC ĐỊNH
        # Kiểm tra xem đã có user nào có role 'admin' chưa
        admin_user = User.query.filter_by(role='admin').first()
        if not admin_user:
            # Tạo mật khẩu hash (admin123)
            hashed_pw = generate_password_hash('admin123', method='pbkdf2:sha256')
            
            new_admin = User(
                username='admin',
                email='admin@gmail.com',
                password_hash=hashed_pw,
                role='admin'
            )
            db.session.add(new_admin)
            db.session.commit()
            print("------------------------------------------")
            print("ĐÃ TẠO TÀI KHOẢN ADMIN MẶC ĐỊNH:")
            print("Username: admin")
            print("Password: admin123")
            print("------------------------------------------")
            
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)